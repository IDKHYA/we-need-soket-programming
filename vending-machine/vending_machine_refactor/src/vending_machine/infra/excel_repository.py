from __future__ import annotations

from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vending_machine.app.dto import DomainEvent
from vending_machine.domain.models import CashInventory, MachineState, Product, Session, VALID_DENOMINATIONS
from vending_machine.infra.file_lock import FileLock


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


class ExcelMachineRepository:
    PRODUCT_SHEET = "products"
    CASH_SHEET = "cash_inventory"
    CONFIG_SHEET = "machine_config"
    SESSION_SHEET = "session_state"
    SALES_LOG_SHEET = "sales_log"
    CASH_LOG_SHEET = "cash_log"
    STOCK_LOG_SHEET = "stock_log"
    AUDIT_LOG_SHEET = "audit_log"

    def __init__(self, workbook_path: str | Path, lock_timeout: float = 5.0):
        self.workbook_path = Path(workbook_path)
        self.lock_timeout = lock_timeout

    def create_template(
        self,
        products: list[Product],
        cash_inventory: CashInventory,
        config: dict[str, str],
        session: Session | None = None,
    ) -> None:
        with self._lock():
            wb = Workbook()
            default = wb.active
            wb.remove(default)

            ws_products = wb.create_sheet(self.PRODUCT_SHEET)
            ws_products.append([
                "product_id", "name", "price", "stock", "max_stock", "active", "image_path", "slot_no"
            ])
            for p in products:
                ws_products.append([
                    p.product_id, p.name, p.price, p.stock, p.max_stock,
                    "Y" if p.active else "N", p.image_path or "", p.slot_no or ""
                ])

            ws_cash = wb.create_sheet(self.CASH_SHEET)
            ws_cash.append(["denomination", "count", "min_keep", "max_capacity", "kind"])
            for denom in sorted(cash_inventory.counts):
                kind = "bill" if denom >= 1000 else "coin"
                ws_cash.append([
                    denom,
                    cash_inventory.counts.get(denom, 0),
                    cash_inventory.min_keep.get(denom, 0),
                    cash_inventory.max_capacity.get(denom, 999999),
                    kind,
                ])

            ws_config = wb.create_sheet(self.CONFIG_SHEET)
            ws_config.append(["key", "value"])
            base_config = {
                "schema_version": "3",
                **config,
            }
            for key, value in base_config.items():
                ws_config.append([key, value])

            ws_session = wb.create_sheet(self.SESSION_SHEET)
            ws_session.append(["key", "value"])
            session = session or Session()
            ws_session.append(["inserted_total", session.inserted_total])
            for denom in VALID_DENOMINATIONS:
                ws_session.append([f"inserted_{denom}", session.inserted_breakdown.get(denom, 0)])

            for sheet_name, headers in [
                (self.SALES_LOG_SHEET, [
                    "sale_id", "sold_at", "product_id", "product_name", "unit_price", "qty",
                    "paid_amount", "change_amount", "result"
                ]),
                (self.CASH_LOG_SHEET, [
                    "cash_event_id", "event_at", "event_type", "denomination", "qty", "amount", "note"
                ]),
                (self.STOCK_LOG_SHEET, [
                    "stock_event_id", "event_at", "product_id", "product_name", "event_type",
                    "before_stock", "change_qty", "after_stock", "note"
                ]),
                (self.AUDIT_LOG_SHEET, [
                    "audit_id", "event_at", "actor", "action", "target", "detail"
                ]),
            ]:
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)

            self._apply_style(wb)
            self._atomic_save(wb)

    def load_state(self) -> MachineState:
        with self._lock():
            wb = load_workbook(self.workbook_path)
            return self._parse_state(wb)

    def load_session(self) -> Session:
        with self._lock():
            wb = load_workbook(self.workbook_path)
            return self._parse_session(wb)

    def save_state(self, state: MachineState) -> None:
        with self._lock():
            wb = load_workbook(self.workbook_path)
            self._write_state(wb, state)
            self._apply_style(wb)
            self._atomic_save(wb)

    def save_session(self, session: Session) -> None:
        with self._lock():
            wb = load_workbook(self.workbook_path)
            self._write_session(wb, session)
            self._apply_style(wb)
            self._atomic_save(wb)

    def append_events(self, events: Iterable[DomainEvent]) -> None:
        events = list(events)
        if not events:
            return
        with self._lock():
            wb = load_workbook(self.workbook_path)
            self._append_events_to_workbook(wb, events)
            self._apply_style(wb)
            self._atomic_save(wb)

    def commit(self, state: MachineState, session: Session, events: Iterable[DomainEvent]) -> None:
        with self._lock():
            wb = load_workbook(self.workbook_path)
            self._write_state(wb, state)
            self._write_session(wb, session)
            self._append_events_to_workbook(wb, list(events))
            self._apply_style(wb)
            self._atomic_save(wb)

    def read_sheet_rows(self, sheet_name: str) -> list[dict[str, object]]:
        with self._lock():
            wb = load_workbook(self.workbook_path, data_only=True)
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            rows: list[dict[str, object]] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(value in (None, "") for value in row):
                    continue
                rows.append({str(header): value for header, value in zip(headers, row)})
            return rows

    def _parse_state(self, wb: Workbook) -> MachineState:
        products_ws = wb[self.PRODUCT_SHEET]
        cash_ws = wb[self.CASH_SHEET]
        config_ws = wb[self.CONFIG_SHEET]

        products = {}
        for row in products_ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            p = Product(
                product_id=str(row[0]),
                name=str(row[1]),
                price=int(row[2]),
                stock=int(row[3]),
                max_stock=int(row[4]),
                active=str(row[5]).upper() == "Y",
                image_path=str(row[6]) if row[6] else None,
                slot_no=int(row[7]) if row[7] not in (None, "") else None,
            )
            products[p.product_id] = p

        cash_counts = {}
        min_keep = {}
        max_capacity = {}
        for row in cash_ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            denom = int(row[0])
            if denom not in VALID_DENOMINATIONS:
                continue
            cash_counts[denom] = int(row[1])
            min_keep[denom] = int(row[2])
            max_capacity[denom] = int(row[3])

        config = {}
        for row in config_ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                config[str(row[0])] = str(row[1]) if row[1] is not None else ""

        return MachineState(
            products=products,
            cash_inventory=CashInventory(
                counts=cash_counts,
                min_keep=min_keep,
                max_capacity=max_capacity,
            ),
            config=config,
        )

    def _parse_session(self, wb: Workbook) -> Session:
        if self.SESSION_SHEET not in wb.sheetnames:
            return Session()
        ws = wb[self.SESSION_SHEET]
        values = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            values[str(row[0])] = int(row[1]) if row[1] not in (None, "") else 0

        breakdown = {}
        for denom in VALID_DENOMINATIONS:
            breakdown[denom] = values.get(f"inserted_{denom}", 0)
        total = values.get("inserted_total", sum(d * q for d, q in breakdown.items()))
        return Session(inserted_total=total, inserted_breakdown=breakdown)

    def _write_state(self, wb: Workbook, state: MachineState) -> None:
        ws = wb[self.PRODUCT_SHEET]
        ws.delete_rows(2, max(0, ws.max_row - 1))
        for product in state.products.values():
            ws.append([
                product.product_id,
                product.name,
                product.price,
                product.stock,
                product.max_stock,
                "Y" if product.active else "N",
                product.image_path or "",
                product.slot_no or "",
            ])

        ws = wb[self.CASH_SHEET]
        ws.delete_rows(2, max(0, ws.max_row - 1))
        for denom in sorted(state.cash_inventory.counts):
            kind = "bill" if denom >= 1000 else "coin"
            ws.append([
                denom,
                state.cash_inventory.counts.get(denom, 0),
                state.cash_inventory.min_keep.get(denom, 0),
                state.cash_inventory.max_capacity.get(denom, 999999),
                kind,
            ])

        ws = wb[self.CONFIG_SHEET]
        ws.delete_rows(2, max(0, ws.max_row - 1))
        for key, value in state.config.items():
            ws.append([key, value])

    def _write_session(self, wb: Workbook, session: Session) -> None:
        ws = wb[self.SESSION_SHEET]
        ws.delete_rows(2, max(0, ws.max_row - 1))
        ws.append(["inserted_total", session.inserted_total])
        for denom in VALID_DENOMINATIONS:
            ws.append([f"inserted_{denom}", session.inserted_breakdown.get(denom, 0)])

    def _append_events_to_workbook(self, wb: Workbook, events: Iterable[DomainEvent]) -> None:
        for event in events:
            ws = wb[event.sheet_name]
            headers = [cell.value for cell in ws[1]]
            row = [event.payload.get(h, "") for h in headers]
            ws.append(row)

    def _apply_style(self, wb: Workbook) -> None:
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            widths: dict[int, int] = {}
            for row in ws.iter_rows():
                for cell in row:
                    value = "" if cell.value is None else str(cell.value)
                    widths[cell.column] = max(widths.get(cell.column, 0), len(value) + 2)
            for col_idx, width in widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 28)

    def _atomic_save(self, workbook: Workbook) -> None:
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        with NamedTemporaryFile(delete=False, suffix=self.workbook_path.suffix, dir=self.workbook_path.parent) as tmp:
            temp_path = Path(tmp.name)
        try:
            workbook.save(temp_path)
            temp_path.replace(self.workbook_path)
        finally:
            temp_path.unlink(missing_ok=True)

    def _lock(self) -> FileLock:
        return FileLock(self.workbook_path, timeout=self.lock_timeout)
